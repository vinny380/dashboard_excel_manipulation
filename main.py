import openpyxl
from dataclasses import dataclass
from io import BytesIO
from typing import List, Optional
import streamlit as st

@dataclass
class ExcelConfig:
    """Configuration settings for Excel processing"""
    HEADER_END_ROW: int = 17
    SUBJECT_NUMBER_CELL: str = 'A12'
    TRIALS_PER_SAMPLE: int = 9
    START_SAMPLE: int = 2
    END_SAMPLE: int = 25
    OUTPUT_FILENAME: str = 'edited_.xlsx'
    COLUMN_HEADERS = [
        'Subject Number', 'trial', 'null', 'condition', 'time',
        'Relationship', 'ControlQ1 Copy 2', 'ControlQ1 Copy - 2 - 2',
        'FirstMoozleProp Copy 13', 'SecondMoozleProp Copy 13',
        'SecondMoozleProp2 Copy 13', 'ChoiceResponse Copy 2',
        'ControlQ2 Copy 2', 'ControlQ2 Copy-2 - 2', 'Choice',
        'SameChoice', 'BeliefType', 'AgeGroup'
    ]

class ExcelProcessor:
    """Handles Excel file processing operations"""
    
    def __init__(self, file_content: BytesIO):
        self.workbook = openpyxl.load_workbook(file_content)
        self.worksheet = self.workbook.active
        self.config = ExcelConfig()

    @staticmethod
    def process_choice(m_column: List, l_column: List, k_column: List) -> List[Optional[str]]:
        """Process choice based on column values"""
        result = []
        for m, l, k in zip(m_column, l_column, k_column):
            if m.value == 'j':
                result.append(l.value)
            elif m.value == 'f':
                result.append(k.value)
            elif m.value == 'd':
                result.append("don't know")
            else:
                result.append(None)
        return result

    @staticmethod
    def process_same_choice(p_column: List, j_column: List) -> List[float]:
        """Process same choice comparison"""
        result = []
        for p, j in zip(p_column, j_column):
            if p.value == j.value:
                result.append(1.0)
            elif p.value == "don't know":
                result.append(0.5)
            else:
                result.append(0.0)
        return result

    @staticmethod
    def get_belief_type(column: List) -> List[str]:
        """Extract belief type from column values"""
        return [str(cell.value)[-1] if cell.value else '' for cell in column]

    def _clean_string(self, value: Optional[str]) -> str:
        """Remove common unwanted patterns from strings"""
        if value is None:
            return ''
        return str(value).replace('.PICT @ :Pictures:', '').replace('[', '').replace(']', '')

    def _get_subject_number(self) -> int:
        """Extract subject number from worksheet"""
        try:
            subject_cell = self.worksheet[self.config.SUBJECT_NUMBER_CELL].value
            if subject_cell:
                return int(subject_cell.replace('Subject Number: ', ''))
        except Exception:
            st.warning("Could not find subject number in cell A12. Using default value 0.")
            return 0

    def process_worksheet(self) -> openpyxl.Workbook:
        """Process the worksheet and return a new workbook"""
        new_workbook = self._initialize_new_workbook()
        new_worksheet = new_workbook.active
        subject_number = self._get_subject_number()

        self._process_samples(new_worksheet, subject_number)
        return new_workbook

    def _initialize_new_workbook(self) -> openpyxl.Workbook:
        """Initialize new workbook with headers"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        for idx, header in enumerate(self.config.COLUMN_HEADERS, 1):
            worksheet.cell(row=1, column=idx, value=header)
            
        return workbook

    def _get_cell_value_safely(self, row: int, col: int) -> str:
        """Safely get cell value with error handling"""
        try:
            value = self.worksheet.cell(row=row, column=col).value
            return self._clean_string(value) if value is not None else ''
        except Exception:
            return ''

    def _process_samples(self, new_worksheet, subject_number: int):
        """Process samples and populate new worksheet"""
        for sample_num in range(self.config.START_SAMPLE, self.config.END_SAMPLE + 1):
            base_row = sample_num
            start_row = (sample_num - 2) * 9 + 1 if sample_num > 2 else 1
            
            # Always process the row, even if some data is missing
            mappings = {
                'Subject Number': (1, lambda: subject_number),
                'trial': (2, lambda: self._get_cell_value_safely(start_row, 3)),
                'condition': (4, lambda: self._get_cell_value_safely(start_row, 6)),
                'time': (5, lambda: self._get_cell_value_safely(start_row + 6, 12)),
                'Relationship': (6, lambda: self._get_cell_value_safely(start_row, 5)),
                'ControlQ1 Copy 2': (7, lambda: self._get_cell_value_safely(start_row + 1, 14)),
                'ControlQ1 Copy - 2 - 2': (8, lambda: self._get_cell_value_safely(start_row + 2, 14)),
                'FirstMoozleProp Copy 13': (9, lambda: self._get_cell_value_safely(start_row + 3, 5)),
                'SecondMoozleProp Copy 13': (10, lambda: self._get_cell_value_safely(start_row + 4, 5)),
                'SecondMoozleProp2 Copy 13': (11, lambda: self._get_cell_value_safely(start_row + 5, 5)),
                'ChoiceResponse Copy 2': (12, lambda: self._get_cell_value_safely(start_row + 6, 14)),
                'ControlQ2 Copy 2': (13, lambda: self._get_cell_value_safely(start_row + 7, 14)),
                'ControlQ2 Copy-2 - 2': (14, lambda: self._get_cell_value_safely(start_row + 8, 14))
            }

            # Apply mappings
            for _, (col_num, value_func) in mappings.items():
                try:
                    value = value_func()
                    new_worksheet.cell(row=base_row, column=col_num, value=value)
                except Exception as e:
                    print(f"Error processing cell: {e}")
                    # Continue processing even if there's an error

def create_streamlit_app():
    """Create and configure Streamlit application"""
    st.title("Language and Cognition Lab🧠📊")
    st.image('https://i.ibb.co/L0hHCFZ/Screenshot-2023-02-22-at-12-57-43-PM.png')
    st.markdown("Automated Excel Editor")
    
    excel_file = st.file_uploader('Upload your excel file')
    
    if st.button('Process Excel'):
        if excel_file:
            try:
                processor = ExcelProcessor(BytesIO(excel_file.read()))
                workbook = processor.process_worksheet()
                
                # Save workbook to BytesIO for download
                output = BytesIO()
                workbook.save(output)
                output.seek(0)
                
                st.download_button(
                    label="Download Updated Excel Workbook",
                    data=output.getvalue(),
                    file_name="workbook.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success("Processing complete! You can now download the processed file.")
                
            except Exception as e:
                st.error(f"An error occurred while processing the file: {str(e)}")
        else:
            st.warning("Please upload an Excel file first.")

    st.write('''
    Note: The following columns will need to be processed after download:
    - SameChoice: =IF(@P:P=@J:J,1,IF(@P:P="don't know",0.5,0))\n
    - BeliefType: =RIGHT(E2, 1)\n
    - AgeGroup: To be processed on lab computer\n
    - Choice: =IF(@M:M="j",L:L,IF(@M:M="f",K:K,IF(@M:M="d","don't know")))\n
    Enjoy!
    ''')

if __name__ == "__main__":
    create_streamlit_app()